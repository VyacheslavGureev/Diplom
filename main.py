# coding: utf-8
import os
import uuid
from PyQt5 import QtCore, QtGui, QtWidgets, QtWebEngineWidgets
import io
import folium
from PyQt5.QtWidgets import QApplication, QWidget, QHBoxLayout, QVBoxLayout, QMenu, QMainWindow, QAction, QToolBar, \
    QTableWidget, QTableWidgetItem, QFileDialog
import openpyxl
import random
from PyQt5.QtGui import QIcon, QStandardItemModel
import psycopg2
import algorithms
from PyQt5.QtCore import QThread, pyqtSignal
from PyQt5.QtWidgets import (QApplication, QDialog,
                             QProgressBar, QPushButton)


LOAD_G_SUCCESS_FLAG = False
LOADING_TERMINATED_FLAG = False
ROUTE = []
SOLVES = {}
COLORS = []

selected_rows = set()
selected_rows_cour_choice = set()
selected_rows_addr_choice = set()
ADDR_DATA_VALID_FLAG = True


def resource_path(relative_path):
    if getattr(sys, 'frozen', False):
        base_path = sys._MEIPASS
    else:
        base_path = os.getcwd()
    return os.path.join(base_path, relative_path)


class Orders():


    def __init__(self):
        self.vehicles = []
        self.finishes = {}
        self.warehouses = []
        self.clients = {}

        self.ready = False

        self.vehiclesMutable = None
        self.finishesMutable = None
        self.warehousesMutable = None
        self.clientsMutable = None


    def data_to_mutable(self):
        self.vehiclesMutable = list(self.vehicles)
        self.finishesMutable = dict(self.finishes)
        self.warehousesMutable = list(self.warehouses)
        self.clientsMutable = dict(self.clients)
        self.ready = True


    def clr_data(self):
        self.vehicles.clear()
        self.finishes.clear()
        self.warehouses.clear()
        self.clients.clear()

        self.vehiclesMutable = None
        self.finishesMutable = None
        self.warehousesMutable = None
        self.clientsMutable = None

        self.ready = False


class Point():


    def __init__(self, **params):
        self.coords = params.get('coords', None)
        self.t = params.get('t', None)
        self.v = params.get('v', None)
        self.T = params.get('T', None)
        self.V = params.get('V', None)
        self.vehicle = params.get('vehicle', None)
        self.type = params.get('type', None)
        self.id = params.get('id', None)
        self.dc = params.get('draw_coords', None)


class TabBar(QtWidgets.QTabBar):


    def tabSizeHint(self, index):
        s = QtWidgets.QTabBar.tabSizeHint(self, index)
        s.transpose()
        return s


    def paintEvent(self, event):
        painter = QtWidgets.QStylePainter(self)
        opt = QtWidgets.QStyleOptionTab()
        for i in range(self.count()):
            self.initStyleOption(opt, i)
            painter.drawControl(QtWidgets.QStyle.CE_TabBarTabShape, opt)
            painter.save()
            s = opt.rect.size()
            s.transpose()
            r = QtCore.QRect(QtCore.QPoint(), s)
            r.moveCenter(opt.rect.center())
            opt.rect = r
            c = self.tabRect(i).center()
            painter.translate(c)
            painter.rotate(90)
            painter.translate(-c)
            painter.drawControl(QtWidgets.QStyle.CE_TabBarTabLabel, opt)
            painter.restore()


class GetAddrData(QThread):

    resSignal = pyqtSignal(list)


    def __init__(self, addrs, algs, cursor, connection, mode):
        QThread.__init__(self)
        self.addrs = addrs
        self.algs = algs
        self.cursor = cursor
        self.connection = connection
        self.mode = mode


    def run(self):
        result = []
        for line in self.addrs:
            for a in line:
                self.cursor.execute("SELECT addr FROM public.addresses WHERE addr='{}'".format(a))
                record = self.cursor.fetchall()
                if (len(record)) == 0:
                    res = []
                    res.append(a)
                    try:
                        coords = self.algs.adr_to_coords(a)
                        res.append(str(coords))
                        res.append(str(self.algs.get_node(coords, self.algs.Gmain_comm_arrs)))
                    except:
                        res.append(None)
                        res.append(None)
                    if res[1] != None:
                        self.cursor.execute("INSERT INTO public.addresses "
                                            "VALUES ('{}', '{}', '{}');".format(res[0], res[1], res[2]))
                        self.connection.commit()
                    result.append(res)
        if self.mode == 'cour':
            self.resSignal.emit(result)


class FindingSolves(QThread):

    solvesSignal = pyqtSignal(dict)


    def __init__(self, algs, orders):
        QThread.__init__(self)
        self.algs = algs
        self.orders = orders


    def run(self):
        r = self.algs.findSolutions(self.orders)
        self.solvesSignal.emit(r)


class FindingRoute(QThread):

    routeSignal = pyqtSignal(list)


    def __init__(self, algs, solves, cursor):
        QThread.__init__(self)
        self.algs = algs
        self.solves = solves
        self.cursor = cursor


    def run(self):
        r = self.algs.find_path(self.solves, self.cursor)
        self.routeSignal.emit(r)


class Progress(QThread):

    progressCountChanged = pyqtSignal(int)


    def __init__(self, *args, **kwargs):
        QThread.__init__(self, *args, **kwargs)
        self.active = True


    def run(self):
        count = 0
        while self.active:
            count += 1
            self.sleep(1)
            self.progressCountChanged.emit(count)


    def stop(self):
        self.active = False


class LoadGraph(QThread):


    def __init__(self, algs, prog):
        QThread.__init__(self)
        self.algs = algs


    def run(self):
        self.algs.load_G_data()


class ProgressDialog(QDialog):


    def __init__(self, prog):
        QDialog.__init__(self)
        self.prog = prog


    def closeEvent(self, event):
        global LOADING_TERMINATED_FLAG
        LOADING_TERMINATED_FLAG = True
        self.prog.stop()
        event.accept()


class Ui_Dialog(object):


    def setupUi(self, Dialog):
        super().__init__()
        Dialog.setObjectName("Dialog")
        Dialog.resize(1600, 900)
        self.tabWidget = QtWidgets.QTabWidget(Dialog)
        self.tabWidget.setGeometry(QtCore.QRect(10, 10, 1550, 870))
        self.tabWidget.setObjectName("tabWidget")

        self.initBD()

        self.planningTab = QtWidgets.QWidget()
        self.planningTab.setObjectName("mapTab")

        self.courTableWidgetPlan = QtWidgets.QTableWidget(self.planningTab)
        self.courTableWidgetPlan.setGeometry(QtCore.QRect(20, 60, 650, 350))
        self.courTableWidgetPlan.setObjectName("courTableWidgetPlan")
        self.courTableWidgetPlan.setColumnCount(7)
        self.courTableWidgetPlan.setRowCount(0)
        self.courTableWidgetPlan.setHorizontalHeaderLabels(
            ["Id", "Фамилия", 'Имя', 'Отчество', "Транспорт", "Начало пути", "Конец пути"])
        self.courTableWidgetPlan.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.courTableWidgetPlan.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
        self.courTableWidgetPlan.resizeColumnsToContents()
        self.addDataCourTableWidgetPlan()
        self.courTableWidgetPlan.selectionModel().selectionChanged.connect(self.onSelectCourPlan)

        self.adrTableWidgetPlan = QtWidgets.QTableWidget(self.planningTab)
        self.adrTableWidgetPlan.setGeometry(QtCore.QRect(20, 470, 650, 350))
        self.adrTableWidgetPlan.setObjectName("adrTableWidget")
        self.adrTableWidgetPlan.setColumnCount(4)
        self.adrTableWidgetPlan.setRowCount(0)
        self.adrTableWidgetPlan.setHorizontalHeaderLabels(
            ["Id", "Товар", 'Адрес заказа', 'Адрес назначения'])
        self.adrTableWidgetPlan.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.adrTableWidgetPlan.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
        self.adrTableWidgetPlan.resizeColumnsToContents()
        self.addDataToAdrTableWidgetPlan()
        self.adrTableWidgetPlan.selectionModel().selectionChanged.connect(self.onSelectAddrPlan)

        self.courLabel = QtWidgets.QLabel(self.planningTab)
        self.courLabel.setGeometry(QtCore.QRect(20, 20, 121, 31))
        self.courLabel.setObjectName("courLabel")

        self.adrLabel = QtWidgets.QLabel(self.planningTab)
        self.adrLabel.setGeometry(QtCore.QRect(20, 430, 151, 31))
        self.adrLabel.setObjectName("adrLabel")

        self.showPointsButton = QtWidgets.QPushButton(self.planningTab)
        self.showPointsButton.setGeometry(QtCore.QRect(720, 20, 124, 30))
        self.showPointsButton.setObjectName("showPointsButton")
        self.showPointsButton.clicked.connect(self.showPoints)

        self.planMoveButton = QtWidgets.QPushButton(self.planningTab)
        self.planMoveButton.setGeometry(QtCore.QRect(864, 20, 124, 30))
        self.planMoveButton.setObjectName("planMoveButton")
        self.planMoveButton.clicked.connect(self.planning)

        self.buildRoutesButton = QtWidgets.QPushButton(self.planningTab)
        self.buildRoutesButton.setGeometry(QtCore.QRect(1008, 20, 124, 30))
        self.buildRoutesButton.setObjectName("buildRoutesButton")
        self.buildRoutesButton.clicked.connect(self.buildRoutes)

        self.saveRouteButton = QtWidgets.QPushButton(self.planningTab)
        self.saveRouteButton.setGeometry(QtCore.QRect(1152, 20, 124, 30))
        self.saveRouteButton.setObjectName("saveRouteButton")
        self.saveRouteButton.clicked.connect(self.savePlan)

        self.loadRouteButton = QtWidgets.QPushButton(self.planningTab)
        self.loadRouteButton.setGeometry(QtCore.QRect(1296, 20, 124, 30))
        self.loadRouteButton.setObjectName("loadRouteButton")
        self.loadRouteButton.clicked.connect(self.loadPlan)

        self.tabWidget.addTab(self.planningTab, "")

        self.createNewFileCour = QAction("Новый файл")
        self.loadFromFileCour = QAction("Загрузить данные из файла")
        self.saveToFileCour = QAction("Сохранить данные в файл")
        self.loadFromDBCour = QAction("Загрузить данные из БД")
        self.saveToDBCour = QAction("Сохранить данные в БД")

        self.couriersMenuBar = QtWidgets.QMenuBar()
        self.couriersMenuBar.setObjectName("menuBar")

        self.fileMenu = QtWidgets.QMenu("Файл")
        self.fileMenu.addAction(self.createNewFileCour)
        self.fileMenu.addAction(self.loadFromFileCour)
        self.fileMenu.addAction(self.saveToFileCour)
        self.fileMenu.addAction(self.loadFromDBCour)
        self.fileMenu.addAction(self.saveToDBCour)
        self.couriersMenuBar.addMenu(self.fileMenu)

        self.createNewFileCour.triggered.connect(self.onCreateNewFileCour)
        self.loadFromFileCour.triggered.connect(self.onLoadFromFileCour)
        self.saveToFileCour.triggered.connect(self.onSaveToFileCour)
        self.loadFromDBCour.triggered.connect(self.onLoadFromDBCour)
        self.saveToDBCour.triggered.connect(self.onSaveToDBCour)

        self.addNewLine = QAction(QIcon(resource_path("resources\plus_green.png")), "Добавить новую запись")
        self.addNewLine.triggered.connect(self.onAddNewLine)

        self.removeSelectedLines = QAction(QIcon(resource_path("resources\minus_red.png")), "Удалить запись")
        self.removeSelectedLines.triggered.connect(self.onRemoveSelectedLines)

        self.deleteSelectedLinesFromDBCour = QAction("Удалить запись из БД")
        self.deleteSelectedLinesFromDBCour.triggered.connect(self.onDeleteSelectedLinesFromDBCour)

        self.couriersMenuBar.addAction(self.addNewLine)
        self.couriersMenuBar.addAction(self.removeSelectedLines)
        self.couriersMenuBar.addAction(self.deleteSelectedLinesFromDBCour)

        self.couriersTableWidget = QtWidgets.QTableWidget(self.couriersMenuBar)
        self.couriersTableWidget.setGeometry(QtCore.QRect(10, 25, 1100, 830))
        self.couriersTableWidget.setObjectName("couriersTableWidget")
        self.couriersTableWidget.setRowCount(0)
        self.couriersTableWidget.setColumnCount(9)
        self.couriersTableWidget.setHorizontalHeaderLabels(["Id", "Фамилия", 'Имя', 'Отчество', "Моб. тел.", 'Название трансп. ср-ва', "Номер трансп. ср-ва", 'Объём (м3)', 'Тоннаж (кг)'])
        self.couriersTableWidget.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.couriersTableWidget.horizontalHeader().setSectionResizeMode(1, QtWidgets.QHeaderView.ResizeToContents)
        self.couriersTableWidget.horizontalHeader().setSectionResizeMode(2, QtWidgets.QHeaderView.ResizeToContents)
        self.couriersTableWidget.horizontalHeader().setSectionResizeMode(3, QtWidgets.QHeaderView.ResizeToContents)
        self.couriersTableWidget.horizontalHeader().setSectionResizeMode(7, QtWidgets.QHeaderView.ResizeToContents)
        self.couriersTableWidget.horizontalHeader().setSectionResizeMode(8, QtWidgets.QHeaderView.ResizeToContents)
        self.couriersTableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
        self.couriersTableWidget.resizeColumnsToContents()
        self.couriersTableWidget.selectionModel().selectionChanged.connect(self.onSelectionChanged)

        self.tabWidget.addTab(self.couriersMenuBar, "")

        self.createNewFileOrd = QAction("Новый файл")
        self.loadFromFileOrd = QAction("Загрузить данные из файла")
        self.saveToFileOrd = QAction("Сохранить данные в файл")
        self.loadFromDBOrd = QAction("Загрузить данные из БД")
        self.saveToDBOrd = QAction("Сохранить данные в БД")

        self.ordersMenuBar = QtWidgets.QMenuBar()
        self.ordersMenuBar.setObjectName("ordersMenuBar")

        self.fileMenuOrd = QtWidgets.QMenu("Файл")
        self.fileMenuOrd.addAction(self.createNewFileOrd)
        self.fileMenuOrd.addAction(self.loadFromFileOrd)
        self.fileMenuOrd.addAction(self.saveToFileOrd)
        self.fileMenuOrd.addAction(self.loadFromDBOrd)
        self.fileMenuOrd.addAction(self.saveToDBOrd)
        self.ordersMenuBar.addMenu(self.fileMenuOrd)

        self.createNewFileOrd.triggered.connect(self.onCreateNewFileOrd)
        self.loadFromFileOrd.triggered.connect(self.onLoadFromFileOrd)
        self.saveToFileOrd.triggered.connect(self.onSaveToFileOrd)
        self.loadFromDBOrd.triggered.connect(self.onLoadFromDBOrd)
        self.saveToDBOrd.triggered.connect(self.onSaveToDBOrd)

        self.addNewLineOrd = QAction(QIcon(resource_path("resources\plus_green.png")), "Добавить новую запись")
        self.addNewLineOrd.triggered.connect(self.onAddNewLineOrd)

        self.removeSelectedLinesOrd = QAction(QIcon(resource_path("resources\minus_red.png")), "Удалить запись")
        self.removeSelectedLinesOrd.triggered.connect(self.onRemoveSelectedLinesOrd)

        self.deleteSelectedLinesFromDBOrd = QAction("Удалить запись из БД")
        self.deleteSelectedLinesFromDBOrd.triggered.connect(self.onDeleteSelectedLinesFromDBOrd)

        self.ordersMenuBar.addAction(self.addNewLineOrd)
        self.ordersMenuBar.addAction(self.removeSelectedLinesOrd)
        self.ordersMenuBar.addAction(self.deleteSelectedLinesFromDBOrd)

        self.ordersTableWidget = QtWidgets.QTableWidget(self.ordersMenuBar)
        self.ordersTableWidget.setGeometry(QtCore.QRect(10, 25, 1100, 830))
        self.ordersTableWidget.setObjectName("couriersTableWidget")

        self.ordersTableWidget.setRowCount(0)
        self.ordersTableWidget.setColumnCount(8)
        self.ordersTableWidget.setHorizontalHeaderLabels(
            ["Id", "Товар", 'Адрес заказа', 'Адрес назначения', "Объём (м3)", 'Тоннаж (кг)', "Время", 'Дата'])
        self.ordersTableWidget.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.Stretch)
        self.ordersTableWidget.horizontalHeader().setSectionResizeMode(4, QtWidgets.QHeaderView.ResizeToContents)
        self.ordersTableWidget.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeToContents)
        self.ordersTableWidget.horizontalHeader().setSectionResizeMode(6, QtWidgets.QHeaderView.ResizeToContents)
        self.ordersTableWidget.horizontalHeader().setSectionResizeMode(7, QtWidgets.QHeaderView.ResizeToContents)
        self.ordersTableWidget.setSizeAdjustPolicy(QtWidgets.QAbstractScrollArea.AdjustToContents)
        self.ordersTableWidget.resizeColumnsToContents()
        self.ordersTableWidget.selectionModel().selectionChanged.connect(self.onSelectionChanged)

        self.tabWidget.setTabBar(TabBar(self.tabWidget))
        self.tabWidget.setTabPosition(self.tabWidget.West)
        self.tabWidget.insertTab(0, self.couriersMenuBar, "")

        self.tabWidget.insertTab(1, self.ordersMenuBar, "")
        self.tabWidget.insertTab(2, self.planningTab, "")

        self.webView = QtWebEngineWidgets.QWebEngineView(self.planningTab)
        self.webView.setObjectName("webView")
        self.webView.setGeometry(QtCore.QRect(720, 60, 700, 760))

        self.orders = Orders()

        self.prog = Progress()
        self.prog.progressCountChanged.connect(self.onProgressCountChanged)
        self.prog.finished.connect(self.deleteProgressDialog)
        self.prog.start()

        self.algs = algorithms.Algorithms()

        self.loadData = LoadGraph(self.algs, self.prog)
        self.loadData.finished.connect(self.loadDataFinish)
        self.loadData.start()

        self.progressDialog = ProgressDialog(self.prog)
        self.progressDialog.setObjectName("ProgressDialog")
        self.progressDialog.setWindowTitle('Загрузка графа')

        self.progress = QProgressBar(self.progressDialog)
        self.progress.setGeometry(0, 0, 300, 25)
        self.progress.setMaximum(5)

        self.map = folium.Map(
            location=[47.183958, 39.737731], zoom_start=10
        )
        data = io.BytesIO()
        self.map.save(data, close_file=False)
        self.webView.setHtml(data.getvalue().decode())

        self.retranslateUi(Dialog)
        self.tabWidget.setCurrentIndex(0)
        self.tabWidget.currentChanged.connect(self.onTabChange)
        QtCore.QMetaObject.connectSlotsByName(Dialog)


    def retranslateUi(self, Dialog):
        _translate = QtCore.QCoreApplication.translate
        Dialog.setWindowTitle(_translate("Dialog", "Планирование грузовых перевозок"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.couriersMenuBar), _translate("Dialog", "Курьеры"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.planningTab), _translate("Dialog", "Планирование"))
        self.tabWidget.setTabText(self.tabWidget.indexOf(self.ordersMenuBar), _translate("Dialog", "Заказы"))

        self.courLabel.setText(_translate("Dialog",
                                          "<html><head/><body><p><span style=\" font-size:14pt;\">Курьеры</span></p></body></html>"))
        self.adrLabel.setText(_translate("Dialog",
                                         "<html><head/><body><p><span style=\" font-size:14pt;\">Адреса доставки</span></p></body></html>"))
        self.showPointsButton.setText(_translate("Dialog", "Спланировать"))
        self.planMoveButton.setText(_translate("Dialog", "Показать план"))
        self.buildRoutesButton.setText(_translate("Dialog", "Построить маршруты"))

        self.saveRouteButton.setText(_translate("Dialog", "Сохранить план"))
        self.loadRouteButton.setText(_translate("Dialog", "Загрузить план"))


    def onTabChange(self):
        if self.tabWidget.currentIndex() == 2:
            self.addDataCourTableWidgetPlan()
            self.addDataToAdrTableWidgetPlan()


    def onProgressCountChanged(self, value):
        self.progress.setValue(value)


    def preproc(self, G, filename):
        self.algs.preproc.G_to_Gform(G, filename)


    def loadDataFinish(self):
        global LOAD_G_SUCCESS_FLAG
        global LOADING_TERMINATED_FLAG
        LOAD_G_SUCCESS_FLAG = True
        self.loadData.deleteLater()
        if LOADING_TERMINATED_FLAG != True:
            self.prog.stop()


    def deleteProgressDialog(self):
        self.prog.deleteLater()
        if LOADING_TERMINATED_FLAG != True:
            self.progressDialog.deleteLater()


    def onSolvesReady(self, result):
        global SOLVES
        global COLORS
        SOLVES = dict(result)
        self.map = folium.Map(location=[47.183958, 39.737731], zoom_start=10)
        self.algs.drawMarkers(self.orders, self.map)
        colors = ['#082567', '#4169E1', '#7851A9', '#CA2C92', '#CA0147',
                  '#E32636', '#F34723', '#3EB489', '#50C878', '#6495ED',
                  '#BD33A4', '#B32821', '#FF4040', '#641349', '#4B0082',
                  '#1CD3A2', '#FF4D00', '#DC143C', '#993366', '#4C5866',
                  '#21421E', '#808000', '#32127A', '#800080']
        for v in result:
            color = random.choice(colors)
            COLORS.append(color)
            points = result[v]
            for i in range(len(points) - 1):
                folium.PolyLine([points[i].dc, points[i+1].dc],
                                color=color,
                                weight=4,
                                opacity=0.8).add_to(self.map)
                self.algs.createArrow(points[i].dc, points[i+1].dc, 20, 60, self.map, color)
        data = io.BytesIO()
        self.map.save(data, close_file=False)
        self.webView.setHtml(data.getvalue().decode())


    def onRouteReady(self, result):
        global ROUTE
        global COLORS
        ROUTE = list(result)
        self.map = folium.Map(location=[47.183958, 39.737731], zoom_start=10)
        self.algs.drawMarkers(self.orders, self.map)
        for i in range(len(result)):
            folium.PolyLine(result[i],
                            color=COLORS[i],
                            weight=4,
                            opacity=0.8).add_to(self.map)
        data = io.BytesIO()
        self.map.save(data, close_file=False)
        self.webView.setHtml(data.getvalue().decode())


    def buildRoutes(self):
        global LOAD_G_SUCCESS_FLAG
        global SOLVES
        if LOAD_G_SUCCESS_FLAG:
            if len(SOLVES) > 0:
                self.findRoute = FindingRoute(algs = self.algs, solves = dict(SOLVES), cursor = self.cursor)
                self.findRoute.routeSignal.connect(self.onRouteReady)
                self.findRoute.finished.connect(self.findRoute.deleteLater)
                self.findRoute.start()
            else:
                self.informDialog = QDialog()
                self.informDialog.setObjectName("InformDialog")
                self.informDialog.resize(450, 90)
                self.informDialogButton = QtWidgets.QPushButton(self.informDialog)
                self.informDialogButton.setGeometry(QtCore.QRect(170, 40, 110, 30))
                self.informDialogButton.setText("Ок")
                self.informDialogButton.setObjectName("informDialogButton")
                self.informDialogButton.clicked.connect(self.acceptInformDialog)
                self.informDialogButtonLabel = QtWidgets.QLabel(self.informDialog)
                self.informDialogButtonLabel.setGeometry(QtCore.QRect(25, 20, 400, 21))
                self.informDialogButtonLabel.setObjectName("informDialogButtonBoxLabel")
                _translate = QtCore.QCoreApplication.translate
                self.informDialog.setWindowTitle(_translate("InformDialog", "Внимание!"))
                self.informDialogButtonLabel.setText(_translate("informDialogButtonBoxLabel",
                                                                "<html><head/><body><p><span style=\" font-size:10pt;\">Движение не спланировано. Пожалуйста, спланируйте движение.</span></p></body></html>"))
                QtCore.QMetaObject.connectSlotsByName(self.informDialog)
                self.informDialog.show()
        else:
            self.informDialog = QDialog()
            self.informDialog.setObjectName("InformDialog")
            self.informDialog.resize(300, 90)
            self.informDialogButton = QtWidgets.QPushButton(self.informDialog)
            self.informDialogButton.setGeometry(QtCore.QRect(70, 40, 160, 30))
            self.informDialogButton.setText("Ок")
            self.informDialogButton.setObjectName("informDialogButton")
            self.informDialogButton.clicked.connect(self.acceptInformDialog)
            self.informDialogButtonLabel = QtWidgets.QLabel(self.informDialog)
            self.informDialogButtonLabel.setGeometry(QtCore.QRect(50, 10, 211, 21))
            self.informDialogButtonLabel.setObjectName("informDialogButtonBoxLabel")
            _translate = QtCore.QCoreApplication.translate
            self.informDialog.setWindowTitle(_translate("InformDialog", "Внимание!"))
            self.informDialogButtonLabel.setText(_translate("informDialogButtonBoxLabel",
                                                            "<html><head/><body><p><span style=\" font-size:12pt;\">Граф ещё не загрузился!</span></p></body></html>"))
            QtCore.QMetaObject.connectSlotsByName(self.informDialog)
            self.informDialog.show()


    def addDataCourTableWidgetPlan(self):
        self.cursor.execute("SELECT courier_id, surname, public.couriers.name, patronymic, public.transport.name "
                            "FROM public.couriers, public.transport "
                            "WHERE courier_id=id_courier")
        record = self.cursor.fetchall()
        self.courTableWidgetPlan.setRowCount(len(record))
        for row in range(self.courTableWidgetPlan.rowCount()):
            item = QTableWidgetItem()
            item.setText(str(record[row][0]))
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsEnabled)
            self.courTableWidgetPlan.setItem(row, 0, item)
            for col in range(1, 5):
                item = QTableWidgetItem()
                item.setText(str(record[row][col]))
                self.courTableWidgetPlan.setItem(row, col, item)
            for col in range(5, 7):
                item = QTableWidgetItem()
                item.setText("")
                self.courTableWidgetPlan.setItem(row, col, item)


    def addDataToAdrTableWidgetPlan(self):
        self.cursor.execute("SELECT order_id, goods, origin_adr, destination_adr "
                            "FROM public.orders")
        record = self.cursor.fetchall()
        self.adrTableWidgetPlan.setRowCount(len(record))
        for row in range(self.adrTableWidgetPlan.rowCount()):
            item = QTableWidgetItem()
            item.setText(str(record[row][0]))
            item.setFlags(QtCore.Qt.ItemFlag.ItemIsEnabled)
            self.adrTableWidgetPlan.setItem(row, 0, item)
            for col in range(1, self.adrTableWidgetPlan.columnCount()):
                item = QTableWidgetItem()
                item.setText(str(record[row][col]))
                self.adrTableWidgetPlan.setItem(row, col, item)


    def onSelectCourPlan(self, selected, deselected):
        global selected_rows_cour_choice
        for ix in deselected.indexes():
            selected_rows_cour_choice.discard(ix.row() + 1)
            for c in range(self.courTableWidgetPlan.columnCount()):
                self.courTableWidgetPlan.item(ix.row(), c).setBackground(QtGui.QColor(255, 255, 255))
        for ix in selected.indexes():
            selected_rows_cour_choice.add(ix.row() + 1)
        print(selected_rows_cour_choice)
        print(selected_rows_addr_choice)
        for r in selected_rows_cour_choice:
            for c in range(self.courTableWidgetPlan.columnCount()):
                self.courTableWidgetPlan.item(r-1, c).setBackground(QtGui.QColor(51, 255, 51))


    def onSelectAddrPlan(self, selected, deselected):
        global selected_rows_addr_choice
        for ix in deselected.indexes():
            selected_rows_addr_choice.discard(ix.row() + 1)
            for c in range(self.adrTableWidgetPlan.columnCount()):
                self.adrTableWidgetPlan.item(ix.row(), c).setBackground(QtGui.QColor(255, 255, 255))
        for ix in selected.indexes():
            selected_rows_addr_choice.add(ix.row() + 1)
        print(selected_rows_cour_choice)
        print(selected_rows_addr_choice)
        for r in selected_rows_addr_choice:
            for c in range(self.adrTableWidgetPlan.columnCount()):
                self.adrTableWidgetPlan.item(r-1, c).setBackground(QtGui.QColor(0, 128, 255))


    def onSelectionChanged(self, selected, deselected):
        global selected_rows
        for ix in deselected.indexes():
            selected_rows.discard(ix.row() + 1)
        for ix in selected.indexes():
            selected_rows.add(ix.row() + 1)
        print(selected_rows)


    def onCreateNewFileCour(self):
        self.couriersTableWidget.setRowCount(0)
        self.couriersTableWidget.setColumnCount(9)


    def onLoadFromDBCour(self):
        self.cursor.execute("SELECT a.courier_id, a.surname, a.name, a.patronymic, "
                            "b.phone, "
                            "c.name, c.num, c.volume, c.tonnage "
                            "FROM public.couriers a "
                            "INNER JOIN public.phones b "
                            "ON a.courier_id=b.id_courier "
                            "INNER JOIN public.transport c "
                            "ON a.courier_id=c.id_courier")
        record = self.cursor.fetchall()
        self.couriersTableWidget.setRowCount(len(record))
        self.couriersTableWidget.setColumnCount(9)
        for row in range(self.couriersTableWidget.rowCount()):
            for col in range(self.couriersTableWidget.columnCount()):
                item = QTableWidgetItem()
                item.setText(str(record[row][col]))
                self.couriersTableWidget.setItem(row, col, item)


    def checkValidDataCour(self, data):
        try:
            for line in range(len(data)):
                for c in range(0, 7):
                    data[line][c] = str(data[line][c])
                for c in range(7, 9):
                    data[line][c] = float(data[line][c])
            return data
        except:
            return 'error!'


    def checkValidDataOrd(self, data):
        try:
            for line in range(len(data)):
                for c in range(0, 4):
                    data[line][c] = str(data[line][c])
                for c in range(4, 6):
                    data[line][c] = float(data[line][c])
                for c in range(6, 8):
                    data[line][c] = str(data[line][c])
            return data
        except:
            return 'error!'


    def onSaveToDBCour(self):
        dataNew = []
        dataUpdate = []
        for row in range(self.couriersTableWidget.rowCount()):
            self.cursor.execute("SELECT courier_id FROM public.couriers WHERE courier_id='{}'".format(
                self.couriersTableWidget.item(row, 0).text()))
            record = self.cursor.fetchall()
            if len(record) == 0:
                line = []
                for column in range(self.couriersTableWidget.columnCount()):
                    line.append(self.couriersTableWidget.item(row, column).text())
                dataNew.append(line)
            else:
                line = []
                for column in range(self.couriersTableWidget.columnCount()):
                    line.append(self.couriersTableWidget.item(row, column).text())
                dataUpdate.append(line)
        if self.checkValidDataCour(dataNew) != 'error!' and self.checkValidDataCour(dataUpdate) != 'error!':
            for i in range(len(dataNew)):
                self.cursor.execute("INSERT INTO public.couriers VALUES ('{}', '{}', '{}', '{}');".format(dataNew[i][1], dataNew[i][2], dataNew[i][3], dataNew[i][0]))
                self.connection.commit()
                self.cursor.execute("INSERT INTO public.phones VALUES ('{}', '{}', '{}');".format(dataNew[i][4], uuid.uuid4(), dataNew[i][0]))
                self.connection.commit()
                self.cursor.execute("INSERT INTO public.transport VALUES ('{}', '{}', '{}', '{}', '{}', '{}');".format(dataNew[i][7], dataNew[i][8],  dataNew[i][5], dataNew[i][6], dataNew[i][0], uuid.uuid4()))
                self.connection.commit()
            for i in range(len(dataUpdate)):
                self.cursor.execute("UPDATE public.couriers "
                                    "SET surname = '{}', name = '{}', patronymic = '{}' "
                                    "WHERE courier_id = '{}'".format(dataUpdate[i][1], dataUpdate[i][2], dataUpdate[i][3], dataUpdate[i][0]))
                self.connection.commit()
                self.cursor.execute("UPDATE public.phones "
                                    "SET phone = '{}' "
                                    "WHERE id_courier = '{}'".format(dataUpdate[i][4], dataUpdate[i][0]))
                self.connection.commit()
                self.cursor.execute("UPDATE public.transport "
                                    "SET volume = {}, tonnage = {}, name = '{}', num = '{}' "
                                    "WHERE id_courier = '{}'".format(dataUpdate[i][7], dataUpdate[i][8], dataUpdate[i][5], dataUpdate[i][6], dataUpdate[i][0]))
                self.connection.commit()
            for row in range(self.couriersTableWidget.rowCount()):
                self.cursor.execute("SELECT courier_id FROM public.couriers WHERE courier_id='{}'".format(
                    self.couriersTableWidget.item(row, 0).text()))
                record = self.cursor.fetchall()
                if len(record) != 0:
                    for column in range(self.couriersTableWidget.columnCount()):
                        self.couriersTableWidget.item(row, column).setBackground(QtGui.QColor(153,255,255))


    def onLoadFromFileCour(self):
        dialog = QFileDialog(self.fileMenu)
        filename, _ = dialog.getOpenFileName(self.fileMenu, "Load From File", "E:\\YandexDisk\\Another\\Diplom\\Proizv\\program\\",  "Excel Files (*.xlsx);;All Files (*)")
        if filename != '':
            wb = openpyxl.load_workbook(filename)
            sh = wb.worksheets[0]
            row_count = sh.max_row - 1
            self.couriersTableWidget.setRowCount(row_count)
            self.couriersTableWidget.setColumnCount(9)
            for row in range(self.couriersTableWidget.rowCount()):
                for col in range(self.couriersTableWidget.columnCount()):
                    item = QTableWidgetItem()
                    item.setText(str(sh.cell(row=row+2, column=col+1).value))
                    self.couriersTableWidget.setItem(row, col, item)


    def onSaveToFileCour(self):
        dialog = QFileDialog(self.fileMenu)
        filename, _ = dialog.getSaveFileName(None, "Save File", "E:\\YandexDisk\\Another\\Diplom\\Proizv\\program\\", "Excel Files (*.xlsx);;All Files (*)")
        if filename != '':
            wb = openpyxl.Workbook()
            sheet = wb.active
            cell = sheet.cell(row=1, column=1)
            cell.value = 'Id'
            cell = sheet.cell(row=1, column=2)
            cell.value = 'Фамилия'
            cell = sheet.cell(row=1, column=3)
            cell.value = 'Имя'
            cell = sheet.cell(row=1, column=4)
            cell.value = 'Отчество'
            cell = sheet.cell(row=1, column=5)
            cell.value = 'Моб. тел.'
            cell = sheet.cell(row=1, column=6)
            cell.value = 'Название трансп. ср-ва'
            cell = sheet.cell(row=1, column=7)
            cell.value = 'Номер трансп. ср-ва'
            cell = sheet.cell(row=1, column=8)
            cell.value = 'Объём (м3)'
            cell = sheet.cell(row=1, column=9)
            cell.value = 'Тоннаж (кг)'
            if self.couriersTableWidget.rowCount() > 0:
                for row in range(self.couriersTableWidget.rowCount()):
                    for col in range(self.couriersTableWidget.columnCount()):
                        sheet.cell(row=row+2, column=col+1).value = self.couriersTableWidget.item(row, col).text()
            wb.save(filename)


    def onAddNewLine(self):
        rowPosition = self.couriersTableWidget.rowCount()
        self.couriersTableWidget.insertRow(rowPosition)
        item = QTableWidgetItem()
        item.setText(str(uuid.uuid4()))
        item.setFlags(QtCore.Qt.ItemFlag.ItemIsEnabled)
        self.couriersTableWidget.setItem(rowPosition, 0, item)
        for col in range(1, self.couriersTableWidget.columnCount()):
            item = QTableWidgetItem()
            item.setText("")
            self.couriersTableWidget.setItem(rowPosition, col, item)


    def onRemoveSelectedLines(self):
        global selected_rows
        if self.couriersTableWidget.rowCount() > 0 and len(selected_rows) > 0:
            rows = []
            for r in selected_rows:
                rows.append(r)
            rows.sort(reverse=True)
            for r in rows:
                self.couriersTableWidget.removeRow(r - 1)


    def onDeleteSelectedLinesFromDBCour(self):
        global selected_rows
        if self.couriersTableWidget.rowCount() > 0 and len(selected_rows) > 0:
            for r in selected_rows:
                self.cursor.execute("SELECT courier_id FROM public.couriers WHERE courier_id='{}'".format(
                    self.couriersTableWidget.item(r-1, 0).text()))
                record = self.cursor.fetchall()
                if len(record) != 0:
                    self.cursor.execute("DELETE FROM public.couriers WHERE courier_id='{}'".format(
                        self.couriersTableWidget.item(r - 1, 0).text()))
                    self.connection.commit()
                    for column in range(self.couriersTableWidget.columnCount()):
                        self.couriersTableWidget.item(r-1, column).setBackground(QtGui.QColor(225,51,51))


    def onCreateNewFileOrd(self):
        self.ordersTableWidget.setRowCount(0)
        self.ordersTableWidget.setColumnCount(9)


    def onLoadFromFileOrd(self):
        dialog = QFileDialog(self.fileMenu)
        filename, _ = dialog.getOpenFileName(self.fileMenu, "Load From File",
                                             "E:\\YandexDisk\\Another\\Diplom\\Proizv\\program\\",
                                             "Excel Files (*.xlsx);;All Files (*)")
        if filename != '':
            wb = openpyxl.load_workbook(filename)
            sh = wb.worksheets[0]
            row_count = sh.max_row - 1
            self.ordersTableWidget.setRowCount(row_count)
            self.ordersTableWidget.setColumnCount(8)
            for row in range(self.ordersTableWidget.rowCount()):
                for col in range(self.ordersTableWidget.columnCount()):
                    item = QTableWidgetItem()
                    item.setText(str(sh.cell(row=row + 2, column=col + 1).value))
                    self.ordersTableWidget.setItem(row, col, item)


    def onSaveToFileOrd(self):
        dialog = QFileDialog(self.fileMenu)
        filename, _ = dialog.getSaveFileName(None, "Save File", "E:\\YandexDisk\\Another\\Diplom\\Proizv\\program\\",
                                             "Excel Files (*.xlsx);;All Files (*)")
        if filename != '':
            wb = openpyxl.Workbook()
            sheet = wb.active
            cell = sheet.cell(row=1, column=1)
            cell.value = 'Id'
            cell = sheet.cell(row=1, column=2)
            cell.value = 'Товар'
            cell = sheet.cell(row=1, column=3)
            cell.value = 'Адрес заказа'
            cell = sheet.cell(row=1, column=4)
            cell.value = 'Адрес назначения'
            cell = sheet.cell(row=1, column=5)
            cell.value = 'Объём (м3)'
            cell = sheet.cell(row=1, column=6)
            cell.value = 'Тоннаж (кг)'
            cell = sheet.cell(row=1, column=7)
            cell.value = 'Время'
            cell = sheet.cell(row=1, column=8)
            cell.value = 'Дата'
            if self.ordersTableWidget.rowCount() > 0:
                for row in range(self.ordersTableWidget.rowCount()):
                    for col in range(self.ordersTableWidget.columnCount()):
                        sheet.cell(row=row + 2, column=col + 1).value = self.ordersTableWidget.item(row, col).text()
            wb.save(filename)


    def onLoadFromDBOrd(self):
        self.cursor.execute("SELECT public.orders.order_id, public.orders.goods, public.orders.origin_adr, public.orders.destination_adr, "
                            "public.orders.volume, public.orders.tonnage, public.orders.time, public.orders.date "
                            "FROM public.orders")
        record = self.cursor.fetchall()
        self.ordersTableWidget.setRowCount(len(record))
        self.ordersTableWidget.setColumnCount(8)
        for row in range(self.ordersTableWidget.rowCount()):
            for col in range(self.ordersTableWidget.columnCount()):
                item = QTableWidgetItem()
                item.setText(str(record[row][col]))
                self.ordersTableWidget.setItem(row, col, item)


    def onSaveToDBOrd(self):
        dataNew = []
        dataUpdate = []
        for row in range(self.ordersTableWidget.rowCount()):
            self.cursor.execute("SELECT order_id FROM public.orders WHERE order_id='{}'".format(
                self.ordersTableWidget.item(row, 0).text()))
            record = self.cursor.fetchall()
            if len(record) == 0:
                line = []
                for column in range(self.ordersTableWidget.columnCount()):
                    line.append(self.ordersTableWidget.item(row, column).text())
                dataNew.append(line)
            else:
                line = []
                for column in range(self.ordersTableWidget.columnCount()):
                    line.append(self.ordersTableWidget.item(row, column).text())
                dataUpdate.append(line)
        if self.checkValidDataOrd(dataNew) != 'error!' and self.checkValidDataOrd(dataUpdate) != 'error!':
            for i in range(len(dataNew)):
                self.cursor.execute(
                    "INSERT INTO public.orders VALUES ('{}', '{}', '{}', '{}', '{}', '{}', '{}', '{}');".format(dataNew[i][2], dataNew[i][3],
                                                                                          dataNew[i][4], dataNew[i][5], dataNew[i][1], dataNew[i][6],
                                                                                          dataNew[i][7], dataNew[i][0]))
                self.connection.commit()
            for i in range(len(dataUpdate)):
                self.cursor.execute(
                    "UPDATE public.orders "
                    "SET goods = '{}', origin_adr = '{}', destination_adr = '{}', volume = {}, tonnage = {}, time = '{}', date = '{}' "
                    "WHERE order_id = '{}'".format(
                        dataUpdate[i][1], dataUpdate[i][2],
                        dataUpdate[i][3], dataUpdate[i][4], dataUpdate[i][5], dataUpdate[i][6],
                        dataUpdate[i][7], dataUpdate[i][0]))
                self.connection.commit()
            for row in range(self.ordersTableWidget.rowCount()):
                self.cursor.execute("SELECT order_id FROM public.orders WHERE order_id='{}'".format(
                    self.ordersTableWidget.item(row, 0).text()))
                record = self.cursor.fetchall()
                if len(record) != 0:
                    for column in range(self.ordersTableWidget.columnCount()):
                        self.ordersTableWidget.item(row, column).setBackground(QtGui.QColor(153, 255, 255))
            addrs = []
            for row in range(self.ordersTableWidget.rowCount()):
                line = []
                line.append(self.ordersTableWidget.item(row, 2).text())
                line.append(self.ordersTableWidget.item(row, 3).text())
                addrs.append(line)
            self.getAddrData = GetAddrData(addrs, self.algs, self.cursor, self.connection, 'ord')
            self.getAddrData.finished.connect(self.getAddrData.deleteLater)
            self.getAddrData.start()


    def prepareAddrs(self, a):
        res = []
        res.append(a)
        try:
            coords = self.algs.adr_to_coords(a)
            res.append(str(coords))
            res.append(str(self.algs.get_node(coords, self.algs.Gmain_comm_arrs)))
        except:
            res.append(None)
            res.append(None)
        return res


    def onAddNewLineOrd(self):
        rowPosition = self.ordersTableWidget.rowCount()
        self.ordersTableWidget.insertRow(rowPosition)
        item = QTableWidgetItem()
        item.setText(str(uuid.uuid4()))
        item.setFlags(QtCore.Qt.ItemFlag.ItemIsEnabled)
        self.ordersTableWidget.setItem(rowPosition, 0, item)
        for col in range(1, self.ordersTableWidget.columnCount()):
            item = QTableWidgetItem()
            item.setText("")
            self.ordersTableWidget.setItem(rowPosition, col, item)


    def onRemoveSelectedLinesOrd(self):
        global selected_rows
        if self.ordersTableWidget.rowCount() > 0 and len(selected_rows) > 0:
            rows = []
            for r in selected_rows:
                rows.append(r)
            rows.sort(reverse=True)
            for r in rows:
                self.ordersTableWidget.removeRow(r - 1)


    def onDeleteSelectedLinesFromDBOrd(self):
        global selected_rows
        if self.ordersTableWidget.rowCount() > 0 and len(selected_rows) > 0:
            for r in selected_rows:
                self.cursor.execute("SELECT order_id FROM public.orders WHERE order_id='{}'".format(
                    self.ordersTableWidget.item(r - 1, 0).text()))
                record = self.cursor.fetchall()
                if len(record) != 0:
                    self.cursor.execute("DELETE FROM public.orders WHERE order_id='{}'".format(
                        self.ordersTableWidget.item(r - 1, 0).text()))
                    self.connection.commit()
                    for column in range(self.ordersTableWidget.columnCount()):
                        self.ordersTableWidget.item(r - 1, column).setBackground(QtGui.QColor(225, 51, 51))


    def savePlan(self):
        dialog = QFileDialog(self.fileMenu)
        filename, _ = dialog.getSaveFileName(None, "Save File", "E:\\YandexDisk\\Another\\Diplom\\Proizv\\program\\",
                                             "Excel Files (*.xlsx);;All Files (*)")
        if filename != '':
            global SOLVES
            wb = openpyxl.Workbook()
            sheet = wb.active
            sheet.cell(row=1, column=1).value = 'Адрес'
            sheet.cell(row=1, column=2).value = 'Latitude'
            sheet.cell(row=1, column=3).value = 'Longitude'
            sheet.cell(row=1, column=4).value = 'Объём (м3)'
            sheet.cell(row=1, column=5).value = 'Тоннаж (кг)'
            row = 2
            for v in SOLVES:
                points = SOLVES[v]
                for p in points:
                    if p.type == 'v':
                        sheet.cell(row=row, column=1).value = f'Начальный адрес: {p.id}'
                        sheet.cell(row=row, column=2).value = f'{p.coords[0]}'
                        sheet.cell(row=row, column=3).value = f'{p.coords[1]}'
                        sheet.cell(row=row, column=4).value = f'{p.V}'
                        sheet.cell(row=row, column=5).value = f'{p.T}'
                    elif p.type == 'w':
                        sheet.cell(row=row, column=1).value = f'Адрес заказа: {p.id}'
                        sheet.cell(row=row, column=2).value = f'{p.coords[0]}'
                        sheet.cell(row=row, column=3).value = f'{p.coords[1]}'
                        sheet.cell(row=row, column=4).value = 'нет'
                        sheet.cell(row=row, column=5).value = 'нет'
                    elif p.type == 'c':
                        sheet.cell(row=row, column=1).value = f'Адрес назначения: {p.id}'
                        sheet.cell(row=row, column=2).value = f'{p.coords[0]}'
                        sheet.cell(row=row, column=3).value = f'{p.coords[1]}'
                        sheet.cell(row=row, column=4).value = f'{p.v}'
                        sheet.cell(row=row, column=5).value = f'{p.t}'
                    elif p.type == 'f':
                        sheet.cell(row=row, column=1).value = f'Конечный адрес: {p.id}'
                        sheet.cell(row=row, column=2).value = f'{p.coords[0]}'
                        sheet.cell(row=row, column=3).value = f'{p.coords[1]}'
                        sheet.cell(row=row, column=4).value = 'нет'
                        sheet.cell(row=row, column=5).value = 'нет'
                    row += 1
                row+=1
            wb.save(filename)


    def loadPlan(self):
        dialog = QFileDialog(self.fileMenu)
        filename, _ = dialog.getOpenFileName(self.fileMenu, "Load From File",
                                             "E:\\YandexDisk\\Another\\Diplom\\Proizv\\program\\",
                                             "Excel Files (*.xlsx);;All Files (*)")
        if filename != '':
            wb = openpyxl.load_workbook(filename)
            sh = wb.active
            sol = {}
            row = 2
            cell = sh.cell(row=row, column=1)
            v = None
            while cell.value != None:
                str = cell.value
                cellval = str.split(":")
                type = cellval[0]
                id = cellval[1]
                if type == 'Начальный адрес':
                    p = Point()
                    p.type = 'v'
                    p.coords = (float(sh.cell(row=row, column=2).value), float(sh.cell(row=row, column=3).value))
                    p.V = float(sh.cell(row=row, column=4).value)
                    p.T = float(sh.cell(row=row, column=5).value)
                    p.id = id
                    v = p
                    sol[p] = [p]
                elif type == 'Адрес заказа':
                    p = Point()
                    p.type = 'w'
                    p.coords = (float(sh.cell(row=row, column=2).value), float(sh.cell(row=row, column=3).value))
                    p.id = id
                    sol[v].append(p)
                elif type == 'Адрес назначения':
                    p = Point()
                    p.type = 'c'
                    p.coords = (float(sh.cell(row=row, column=2).value), float(sh.cell(row=row, column=3).value))
                    p.v = float(sh.cell(row=row, column=4).value)
                    p.t = float(sh.cell(row=row, column=5).value)
                    p.id = id
                    sol[v].append(p)
                elif type == 'Конечный адрес':
                    p = Point()
                    p.type = 'f'
                    p.coords = (float(sh.cell(row=row, column=2).value), float(sh.cell(row=row, column=3).value))
                    p.id = id
                    sol[v].append(p)
                if sh.cell(row=row+1, column=1).value == None:
                    if sh.cell(row=row+2, column=1).value != None:
                        row+=2
                        cell = sh.cell(row=row, column=1)
                    else:
                        break
                else:
                    row += 1
                    cell = sh.cell(row=row, column=1)


    def onAddrDataReady(self):
        global selected_rows_cour_choice
        global selected_rows_addr_choice
        self.orders.clr_data()
        vehicles = self.orders.vehicles
        finishes = self.orders.finishes
        warehouses = self.orders.warehouses
        clients = self.orders.clients
        for r in selected_rows_cour_choice:
            coords = self.getCoordsFromAddr(self.courTableWidgetPlan.item(r-1, 5).text())

            self.cursor.execute("SELECT volume, tonnage "
                                "FROM public.transport "
                                "WHERE id_courier='{}'".format(self.courTableWidgetPlan.item(r - 1, 0).text()))
            record = self.cursor.fetchall()
            V = float(record[0][0])
            T = float(record[0][1])
            pv = Point(coords=coords, T=T, V=V, id=self.courTableWidgetPlan.item(r - 1, 5).text(), type='v')
            vehicles.append(pv)
            coords = self.getCoordsFromAddr(self.courTableWidgetPlan.item(r-1, 6).text())
            pf = Point(coords=coords, id=self.courTableWidgetPlan.item(r - 1, 6).text(), type='f')
            finishes[pv] = pf
        ordsid = []
        for r in selected_rows_addr_choice:
            ordsid.append(self.adrTableWidgetPlan.item(r - 1, 0).text())
        self.cursor.execute("SELECT origin_adr "
                            "FROM (SELECT origin_adr "
                            "FROM public.orders WHERE order_id IN %s) AS foo "
                            "GROUP BY origin_adr", [tuple(ordsid)])
        record = self.cursor.fetchall()
        for w in record:
            coords = self.getCoordsFromAddr(w[0])
            warehouse = Point(coords=coords, type='w', id = w[0])
            warehouses.append(warehouse)
            self.cursor.execute("SELECT destination_adr, SUM(volume), SUM(tonnage) FROM "
                                "(SELECT * FROM public.orders "
                                "WHERE origin_adr = '{}' AND "
                                "order_id IN %s) AS foo "
                                "GROUP BY destination_adr".format(w[0]), [tuple(ordsid)])
            rec = self.cursor.fetchall()
            c = []
            for i in rec:
                coords = self.getCoordsFromAddr(i[0])
                c.append(Point(coords=coords, t=i[2], v=i[1], type='c', id = i[0]))
            clients[warehouse] = c
        self.orders.data_to_mutable()
        self.clearMap()
        self.algs.drawMarkers(self.orders, self.map)
        for v in finishes:
            folium.PolyLine([v.dc, finishes[v].dc],
                            color='#ff6800',
                            weight=4,
                            opacity=1).add_to(self.map)
            self.algs.createArrow(v.dc, finishes[v].dc, 20, 60, self.map, '#ff6800')
        for w in clients:
            c = clients[w]
            for i in c:
                folium.PolyLine([w.dc, i.dc],
                                color='#1959d1',
                                weight=4,
                                opacity=1).add_to(self.map)
                self.algs.createArrow(w.dc, i.dc, 20, 60, self.map, '#1959d1')
        data = io.BytesIO()
        self.map.save(data, close_file=False)
        self.webView.setHtml(data.getvalue().decode())
        self.getAddrData.deleteLater()


    def getCoordsFromAddr(self, adr):
        self.cursor.execute("SELECT coords "
                            "FROM public.addresses "
                            "WHERE addr='{}'".format(adr))
        rec = self.cursor.fetchall()
        c = rec[0][0]
        c = c.replace('(', '')
        c = c.replace(')', '')
        c = c.replace(', ', ' ').split()
        coords = (float(c[0]), float(c[1]))
        return coords


    def clearMap(self):
        self.map = folium.Map(
            location=[47.183958, 39.737731], zoom_start=9
        )
        data = io.BytesIO()
        self.map.save(data, close_file=False)
        self.webView.setHtml(data.getvalue().decode())


    def showPoints(self):
        global selected_rows_cour_choice
        global selected_rows_addr_choice
        if len(selected_rows_cour_choice) == 0:
            for row in range(self.courTableWidgetPlan.rowCount()):
                selected_rows_cour_choice.add(row)
        if len(selected_rows_addr_choice) == 0:
            for row in range(self.adrTableWidgetPlan.rowCount()):
                selected_rows_addr_choice.add(row)
        addrs = []
        for row in range(self.courTableWidgetPlan.rowCount()):
            line = []
            line.append(self.courTableWidgetPlan.item(row, 5).text())
            line.append(self.courTableWidgetPlan.item(row, 6).text())
            addrs.append(line)
        self.getAddrData = GetAddrData(addrs, self.algs, self.cursor, self.connection, 'cour')
        self.getAddrData.finished.connect(self.onAddrDataReady)
        self.getAddrData.start()


    def planning(self):
        global LOAD_G_SUCCESS_FLAG
        if LOAD_G_SUCCESS_FLAG:
            if self.orders.ready:
                self.findSolves = FindingSolves(self.algs, self.orders)
                self.findSolves.solvesSignal.connect(self.onSolvesReady)
                self.findSolves.finished.connect(self.findSolves.deleteLater)
                self.findSolves.start()
            else:
                self.informDialog = QDialog()
                self.informDialog.setObjectName("InformDialog")
                self.informDialog.resize(300, 90)

                self.informDialogButton = QtWidgets.QPushButton(self.informDialog)
                self.informDialogButton.setGeometry(QtCore.QRect(70, 40, 160, 30))
                self.informDialogButton.setText("Ок")
                self.informDialogButton.setObjectName("informDialogButton")
                self.informDialogButton.clicked.connect(self.acceptInformDialog)

                self.informDialogButtonLabel = QtWidgets.QLabel(self.informDialog)
                self.informDialogButtonLabel.setGeometry(QtCore.QRect(50, 10, 211, 21))
                self.informDialogButtonLabel.setObjectName("informDialogButtonBoxLabel")
                _translate = QtCore.QCoreApplication.translate
                self.informDialog.setWindowTitle(_translate("InformDialog", "Внимание!"))
                self.informDialogButtonLabel.setText(_translate("informDialogButtonBoxLabel",
                                                                "<html><head/><body><p><span style=\" font-size:12pt;\">Граф ещё не загрузился</span></p></body></html>"))
                QtCore.QMetaObject.connectSlotsByName(self.informDialog)
                self.informDialog.show()
        else:
            self.informDialog = QDialog()
            self.informDialog.setObjectName("InformDialog")
            self.informDialog.resize(300, 90)

            self.informDialogButton = QtWidgets.QPushButton(self.informDialog)
            self.informDialogButton.setGeometry(QtCore.QRect(70, 40, 160, 30))
            self.informDialogButton.setText("Ок")
            self.informDialogButton.setObjectName("informDialogButton")
            self.informDialogButton.clicked.connect(self.acceptInformDialog)

            self.informDialogButtonLabel = QtWidgets.QLabel(self.informDialog)
            self.informDialogButtonLabel.setGeometry(QtCore.QRect(50, 10, 211, 21))
            self.informDialogButtonLabel.setObjectName("informDialogButtonBoxLabel")
            _translate = QtCore.QCoreApplication.translate
            self.informDialog.setWindowTitle(_translate("InformDialog", "Внимание!"))
            self.informDialogButtonLabel.setText(_translate("informDialogButtonBoxLabel", "<html><head/><body><p><span style=\" font-size:12pt;\">Граф ещё не загрузился</span></p></body></html>"))
            QtCore.QMetaObject.connectSlotsByName(self.informDialog)
            self.informDialog.show()


    def acceptInformDialog(self):
        self.informDialog.deleteLater()


    def initBD(self):
        self.connection = psycopg2.connect(database="routing_service_db", user="postgres", password="129034qQ", host="localhost", port=1290)
        self.cursor = self.connection.cursor()


if __name__ == "__main__":
    import sys
    app = QtWidgets.QApplication(sys.argv)
    Dialog = QtWidgets.QDialog()
    ui = Ui_Dialog()
    ui.setupUi(Dialog)
    Dialog.show()
    ui.progressDialog.show()
    app.exec_()